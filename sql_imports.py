from fastapi import FastAPI, HTTPException, UploadFile, File, Request
from fastapi.responses import HTMLResponse
from fastapi.templating import Jinja2Templates
import mysql.connector
from openpyxl import load_workbook
import datetime
import os
import shutil
import logging
from typing import Dict, Any
import uvicorn  

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Setup FastAPI app
app = FastAPI()

# Setup templates only
templates = Jinja2Templates(directory="templates")

# Create uploads directory
UPLOAD_DIR = os.path.join(os.path.dirname(__file__), 'uploads')
os.makedirs(UPLOAD_DIR, exist_ok=True)

# Database configuration
DB_CONFIG = {
    "host": "127.0.0.1",
    "port": 3306,
    "user": "hadi",
    "password": "123",
    "database": "Task"
}

@app.get("/", response_class=HTMLResponse)
async def index(request: Request):  # Add Request import and parameter
    return templates.TemplateResponse(
        "index.html", 
        {
            "request": request,
            "message": None
        }
    )

@app.post("/process")
async def process_file(file: UploadFile = File(...)):
    try:
        # Save uploaded file
        file_path = os.path.join(UPLOAD_DIR, "uploaded_file.xlsx")
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        # Process the file
        success, message = process_excel_data(file_path)
        
        if success:
            return {"status": "success", "message": message}
        else:
            raise HTTPException(status_code=500, detail=message)
            
    except Exception as e:
        logger.error(f"Error processing file: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

def sanitize_identifier(name: str) -> str:
    """Clean up column and table names"""
    if not name or not isinstance(name, str) or name.strip() == '':
        return None  # Return None for empty values
    
    # Clean the name
    sanitized = str(name).strip()
    sanitized = ''.join(c if c.isalnum() else '_' for c in sanitized)
    sanitized = sanitized.replace(' ', '_')
    
    # Ensure valid SQL identifier
    if sanitized[0].isdigit():
        sanitized = 'col_' + sanitized
        
    return sanitized

def process_excel_data(file_path: str) -> tuple[bool, str]:
    conn = None
    try:
        workbook = load_workbook(filename=file_path, data_only=True)
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor()
        
        tables_created = []
        rows_inserted = 0
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Clean table name
            table_name = sanitize_identifier(sheet_name.upper())
            if not table_name:
                logger.warning(f"Skipping sheet with invalid name: {sheet_name}")
                continue
                
            # Get headers from first row
            raw_headers = [cell.value for cell in worksheet[1]]
            
            # Clean and validate headers
            headers = []
            for i, header in enumerate(raw_headers):
                clean_header = sanitize_identifier(header)
                if not clean_header:
                    clean_header = f"column_{i+1}"
                headers.append(clean_header)
            
            # Create table
            create_sql = f"""
            CREATE TABLE IF NOT EXISTS `{table_name}` (
                {', '.join(f'`{h}` VARCHAR(255)' for h in headers)}
            )"""
            
            logger.info(f"Creating table {table_name}")
            logger.debug(f"SQL: {create_sql}")
            cursor.execute(create_sql)
            
            # Insert data
            insert_sql = f"""
            INSERT INTO `{table_name}` 
            (`{'`, `'.join(headers)}`) 
            VALUES ({', '.join(['%s'] * len(headers))})
            """
            
            # Process rows
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):  # Skip empty rows
                    cursor.execute(insert_sql, row[:len(headers)])
                    rows_inserted += 1
            
            conn.commit()
            tables_created.append(table_name)
            
        return True, f"Created {len(tables_created)} tables with {rows_inserted} rows"
        
    except Exception as e:
        logger.error(f"Database error: {str(e)}")
        if conn:
            conn.rollback()
        return False, str(e)
        
    finally:
        if conn:
            conn.close()

if __name__ == "__main__":
    try:
        logger.info("Starting server...")
        logger.info("Server will be available at http://localhost:8000")
        uvicorn.run(
            "sql_imports:app",  # Changed from app to string reference
            host="0.0.0.0", 
            port=8000, 
            reload=True,
            log_level="info"
        )
    except Exception as e:
        logger.error(f"Server startup failed: {e}")
        raise